# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Helper function to check if a value is considered missing
MISSING_VALUES = ["", "nan", "nan ", "-", "#n/d", "n/d", "#¡valor!"]
def is_missing(val):
    if val is None:
        return True
    if isinstance(val, float):
        return pd.isna(val)
    return str(val).strip().lower() in MISSING_VALUES

# Helper function to format cell comments with bold titles and italic values
# Each field is shown on a new line, numeric values with 3 significant digits, section titles in bold
# openpyxl comments do not support rich text, so we use Markdown-like formatting for clarity

def format_comment(dictionary, title=None, indent=0, max_indent=2):
    import numbers
    if not dictionary:
        return ''
    lines = []
    prefix = '    ' * indent
    # Separador visual para la sección
    if title and indent == 0:
        lines.append(f"=== {title.upper()} ===")
    for k, v in dictionary.items():
        if k == "Detalle imputación":
            continue  # Skip this field
        # Si es un subdiccionario y no estamos en la última anidación
        if isinstance(v, dict) and indent < max_indent:
            lines.append(f"{prefix}{k}:")
            sub_comment = format_comment(v, None, indent=indent+1, max_indent=max_indent)
            if sub_comment:
                lines.append(sub_comment)
        # Si es un subdiccionario en la última anidación, mostrar como lista en una sola línea
        elif isinstance(v, dict) and indent >= max_indent:
            sub_items = []
            for subk, subv in v.items():
                if isinstance(subv, float):
                    value = f"{subv:.3g}"
                    if 'e' in value or 'E' in value:
                        value = f"{float(subv):.3f}"
                elif isinstance(subv, numbers.Number):
                    value = f"{float(subv):.3g}"
                    if 'e' in value or 'E' in value:
                        value = f"{float(subv):.3f}"
                elif hasattr(subv, 'item') and callable(getattr(subv, 'item', None)):
                    try:
                        val = subv.item()
                        value = f"{float(val):.3g}"
                        if 'e' in value or 'E' in value:
                            value = f"{float(val):.3f}"
                    except Exception:
                        value = str(subv)
                else:
                    value = str(subv)
                sub_items.append(f"{subk}: {value}")
            lines.append(f"{prefix}{k}: [{'; '.join(sub_items)}]")
        # Si es una lista de dicts, imprimir cada uno en nueva línea
        elif isinstance(v, list) and v and all(isinstance(i, dict) for i in v):
            lines.append(f"{prefix}{k}:")
            for i, subdict in enumerate(v):
                lines.append(f"{prefix}  - Item {i+1}:")
                sub_comment = format_comment(subdict, None, indent=indent+2, max_indent=max_indent)
                if sub_comment:
                    lines.append(sub_comment)
        # Si es una lista de valores simples, imprimir todos en una sola línea
        elif isinstance(v, list):
            value_list = []
            for item in v:
                if isinstance(item, float):
                    value = f"{item:.3g}"
                    if 'e' in value or 'E' in value:
                        value = f"{float(item):.3f}"
                elif isinstance(item, numbers.Number):
                    value = f"{float(item):.3g}"
                    if 'e' in value or 'E' in value:
                        value = f"{float(item):.3f}"
                elif hasattr(item, 'item') and callable(getattr(item, 'item', None)):
                    try:
                        val = item.item()
                        value = f"{float(val):.3g}"
                        if 'e' in value or 'E' in value:
                            value = f"{float(val):.3f}"
                    except Exception:
                        value = str(item)
                else:
                    value = str(item)
                value_list.append(value)
            lines.append(f"{prefix}{k}: [{', '.join(value_list)}]")
        # Si es un valor numérico, formatear a 3 cifras significativas y evitar notación científica
        elif isinstance(v, float):
            value = f"{v:.3g}"
            if 'e' in value or 'E' in value:
                value = f"{float(v):.3f}"
            lines.append(f"{prefix}{k}:   {value}")
        elif isinstance(v, numbers.Number):
            value = f"{float(v):.3g}"
            if 'e' in value or 'E' in value:
                value = f"{float(v):.3f}"
            lines.append(f"{prefix}{k}:   {value}")
        elif hasattr(v, 'item') and callable(getattr(v, 'item', None)):
            try:
                val = v.item()
                value = f"{float(val):.3g}"
                if 'e' in value or 'E' in value:
                    value = f"{float(val):.3f}"
            except Exception:
                value = str(v)
            lines.append(f"{prefix}{k}:   {value}")
        else:
            lines.append(f"{prefix}{k}:   {str(v)}")
    if indent == 0:
        lines.append("")
    return '\n'.join(lines)


def exportar_excel_con_imputaciones(source_file, df_processed, details_for_excel, output_file=r"C:\Users\delpi\OneDrive\Tesis\ADRpy-VTOL\ADRpy\analisis\Results\Datos_imputados.xlsx"):
    """
    Exports the processed DataFrame to an Excel file, preserving the original format.
    Adds colors and comments to cells imputed by similarity, correlation, or both, including full details for each method used.

    :param source_file: Path to the original Excel file.
    :param output_file: Path to the output Excel file.
    :param df_processed: DataFrame with the imputed values.
    :param details_for_excel: List of dicts with details for each imputation (final, similarity, correlation).
    """
    try:
        if not details_for_excel:
            print("No imputations to export.")
            return

        print(f"=== Exporting data to file: {output_file} ===")
        wb = load_workbook(source_file)
        # Robust check for active sheet
        if wb.sheetnames:
            ws = wb.active
            if ws is None:
                print("Error: Could not get the active sheet from the Excel file.")
                return
        else:
            print(f"Error: The file '{source_file}' contains no sheets.")
            return

        # Define fill colors for each imputation method
        color_similarity = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        color_correlation = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        color_weighted = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")    # Blue
        color_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")      # Orange

        # Build a quick-access dictionary by cell
        details_dict = {(d["Parámetro"], d["Aeronave"]): d for d in details_for_excel}

        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if ws is None or cell is None or cell.column is None or cell.row is None:
                    continue
                parameter = ws.cell(row=1, column=cell.column).value
                aircraft = ws.cell(row=cell.row, column=1).value
                key = (parameter, aircraft)
                if key in details_dict:
                    detail = details_dict[key]
                    imputed_value = df_processed.at[aircraft, parameter]
                    cell.value = imputed_value
                    # Validity check for imputed value
                    valid_sim = detail["similitud"] and not is_missing(detail["similitud"].get("Valor imputado", None))
                    valid_corr = detail["correlacion"] and not is_missing(detail["correlacion"].get("Valor imputado", None))
                    valid_weighted = detail["final"] and not is_missing(detail["final"].get("Valor imputado", None)) and valid_sim and valid_corr
                    # Color logic
                    if valid_weighted:
                        cell.fill = color_weighted
                    elif valid_sim:
                        cell.fill = color_similarity
                    elif valid_corr:
                        cell.fill = color_correlation
                    elif detail["similitud"] or detail["correlacion"]:
                        # Evaluated but no valid value
                        cell.fill = color_orange
                    # Build clean, ordered comment
                    comment = ''
                    if detail["final"]:
                        comment += format_comment(detail["final"], "IMPUTED VALUE")
                    if detail["similitud"]:
                        sim_comment = format_comment(detail["similitud"], "SIMILARITY DETAILS")
                        if sim_comment:
                            comment += "\n" + sim_comment
                    if detail["correlacion"]:
                        corr_comment = format_comment(detail["correlacion"], "CORRELATION DETAILS")
                        if corr_comment:
                            comment += "\n" + corr_comment
                    if comment:
                        cell.comment = Comment(comment, "System")
        wb.save(output_file)
        print(f"Export completed. File saved as '{output_file}'.")
    except FileNotFoundError:
        print(f"Error: File '{source_file}' or {output_file} not found.")
    except Exception as e:
        print(f"Error processing the file: {e}")
